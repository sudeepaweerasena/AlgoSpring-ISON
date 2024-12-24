import openpyxl
import pandas as pd
from datetime import datetime

# File paths and sheet names
OUTPUT_FILE_PATH = "C:\\Users\\sudeepa.w\\Documents\\GitHub\\AlgoSpring-ISON\\MemberUpload.xlsx"
OUTPUT_SHEET_NAME = "loader"
INPUT_FILE_PATH = "C:\\Users\\sudeepa.w\\Documents\\GitHub\\AlgoSpring-ISON\\CensusData.xlsx"
INPUT_SHEET_NAME = "Sheet1"

def nlg_transfer_medical_data():
    try:
        input_df = pd.read_excel(INPUT_FILE_PATH, sheet_name=INPUT_SHEET_NAME)
        print("Initial Input Data:")
        print(input_df.head())

    except FileNotFoundError:
        print(f"Error: The file {INPUT_FILE_PATH} does not exist.")
        return
    except ValueError:
        print(f"Error: The sheet {INPUT_SHEET_NAME} does not exist in {INPUT_FILE_PATH}.")
        return

    # Replace 'Principal' with 'Employee' in the 'Relation' column
    input_df['Relation'] = input_df['Relation'].replace('Principal', 'Employee')
    print("After Replacing 'Principal' with 'Employee':")
    print(input_df.head())

    column_mapping = {
        "Relation": "Relation",
        "Gender": "Gender",
        "DOB": "DOB",
        "Category": "Category",
        "Marital status": "Marital Status"
    }

    category_mapping = {'A': "Cat A", 'B': "Cat B", 'C': "Cat C"}

    input_df = input_df.rename(columns=column_mapping)

    try:
        output_wb = openpyxl.load_workbook(OUTPUT_FILE_PATH)
        output_ws = output_wb[OUTPUT_SHEET_NAME]
    except FileNotFoundError:
        print(f"Error: The file {OUTPUT_FILE_PATH} does not exist.")
        return
    except KeyError:
        print(f"Error: The sheet {OUTPUT_SHEET_NAME} does not exist in {OUTPUT_FILE_PATH}.")
        return

    for index, row in input_df.iterrows():
        dob = row.get("DOB", "")
        if pd.notnull(dob):
            try:
                dob_converted = pd.to_datetime(dob, errors='coerce', format="%m/%d/%Y")
                if pd.notnull(dob_converted):
                    formatted_dob = dob_converted.strftime("%d/%m/%Y")  # Format changed to mm/dd/yyyy
                else:
                    formatted_dob = "Invalid DOB"
            except ValueError:
                formatted_dob = "Invalid DOB"
            print(f"Processed DOB for {row.get('Full Name', '')}: {formatted_dob}")  # Debugging line
        else:
            formatted_dob = "Invalid DOB"
        output_ws.cell(row=index + 2, column=3).value = formatted_dob

        output_ws.cell(row=index + 2, column=7).value = row.get("Marital Status", "")
        output_ws.cell(row=index + 2, column=2).value = row.get("Gender", "")
        relation = row.get("Relation", "")
        output_ws.cell(row=index + 2, column=1).value = relation
        output_ws.cell(row=index + 2, column=5).value = "DXB"  # Visa Location
        output_ws.cell(row=index + 2, column=4).value = "Enhanced"  # Salary Type
        category_letter = row.get("Category", "")
        category_value = category_mapping.get(category_letter, "Unknown")
        output_ws.cell(row=index + 2, column=6).value = category_value

    output_wb.save(OUTPUT_FILE_PATH)
    print(f"Data successfully written to {OUTPUT_FILE_PATH}")

# # Call the function to execute
# nlg_transfer_medical_data()
